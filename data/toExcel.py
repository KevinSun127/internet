from openpyxl import Workbook, load_workbook
import itertools


def createWb(filename):
    wb = Workbook()
    if(filename.find('.') == -1):
        filename+=".xlsx"
    wb.save(filename)


def createSheet(filename, sheetName, idx):
    if(filename.find('.') == -1):
        filename+=".xlsx"
    wb = load_workbook(filename)
    wb.create_sheet(sheetName, idx)
    wb.save(filename)

def addData(filename, sheetName, dataList):
    wb = load_workbook(filename)
    active = wb[sheetName]
    for row in dataList:
        active.append(row)
    wb.save(filename)

